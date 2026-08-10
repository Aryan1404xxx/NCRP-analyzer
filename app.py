import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
import io

st.set_page_config(page_title="NCRP Daily Analyzer", page_icon="🏛️", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp {
    background: linear-gradient(-45deg, #0f0c29, #302b63, #24243e, #1a1a2e);
    background-size: 400% 400%;
    animation: gradientShift 12s ease infinite;
}
@keyframes gradientShift {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}
.main-header {
    background: linear-gradient(135deg, #1a3a5c, #2563a8);
    padding: 2rem 2.5rem; border-radius: 16px; margin-bottom: 1.5rem;
}
.main-header h1 { color: white; font-size: 2rem; font-weight: 700; margin: 0; }
.main-header p  { color: rgba(255,255,255,0.75); margin: 0.3rem 0 0; font-size: 0.95rem; }
.metric-card {
    background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.15);
    border-radius: 12px; padding: 1rem 1.2rem; text-align: center;
    backdrop-filter: blur(10px);
}
.metric-num   { font-size: 1.8rem; font-weight: 700; color: white; }
.metric-label { font-size: 0.75rem; color: rgba(255,255,255,0.6); margin-top: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
.section-header {
    font-size: 1rem; font-weight: 600; color: #1a3a5c;
    border-left: 4px solid #2563a8; padding-left: 10px; margin: 1.5rem 0 0.75rem;
}
section[data-testid="stSidebar"] { background: #1a3a5c; }
section[data-testid="stSidebar"] * { color: white !important; }
.stButton > button {
    background: #2563a8; color: white; border: none;
    border-radius: 8px; font-weight: 600; padding: 0.5rem 1.5rem;
}
.stTabs [aria-selected="true"] { color: #2563a8 !important; border-bottom: 3px solid #2563a8 !important; font-weight: 600; }
div[data-testid="stMetricValue"] { color: #1a3a5c !important; font-weight: 700 !important; }
h1, h2, h3, h4, p, label, .stMarkdown { color: white !important; }
div[data-testid="stMetricValue"] { color: white !important; font-weight: 700 !important; }
div[data-testid="stMetricLabel"] { color: rgba(255,255,255,0.6) !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🏛️ NCRP Daily Transaction Analyzer</h1>
    <p>Upload daily NCRP Excel file — get instant breakdown of ACKs, accounts, layers, actions, states and more</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("## ℹ️ How to Use")
    st.markdown("1️⃣ Upload the daily Excel file")
    st.markdown("2️⃣ Select the sheet with transactions")
    st.markdown("3️⃣ View analysis across all tabs")
    st.markdown("4️⃣ Download full Excel report")
    st.markdown("---")
    st.markdown("**📊 Report Includes**")
    st.markdown("• Daily Summary")
    st.markdown("• Unique ACK Numbers")
    st.markdown("• Unique Account Numbers")
    st.markdown("• Layer-wise Breakdown")
    st.markdown("• Action Breakdown")
    st.markdown("• State-wise Breakdown")
    st.markdown("• Bank Breakdown")
    st.markdown("• Officer Breakdown")
    st.markdown("---")
    st.markdown("**Developed by:** Aryan Sinha")

@st.cache_data
def load_data(file_bytes, sheet_name):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]
    df['Transaction Amount'] = pd.to_numeric(df.get('Transaction Amount', 0), errors='coerce').fillna(0)
    df['Disputed Amount']    = pd.to_numeric(df.get('Disputed Amount', 0),    errors='coerce').fillna(0)
    df['Layers']             = pd.to_numeric(df.get('Layers', 0),             errors='coerce').fillna(0)
    if 'Acknowledgement No.' in df.columns:
        df['Acknowledgement No.'] = df['Acknowledgement No.'].astype(str).str.strip()
    if 'Account Number' in df.columns:
        df['Account Number'] = df['Account Number'].astype(str).str.strip()
    return df

def build_excel(df, date_label):
    wb = openpyxl.Workbook()

    NAVY   = PatternFill("solid", fgColor="1a3a5c")
    BLUE   = PatternFill("solid", fgColor="2563a8")
    GREEN  = PatternFill("solid", fgColor="059669")
    RED    = PatternFill("solid", fgColor="DC2626")
    ORANGE = PatternFill("solid", fgColor="D97706")
    PURPLE = PatternFill("solid", fgColor="7C3AED")
    ALT1   = PatternFill("solid", fgColor="EFF6FF")
    WHITE  = PatternFill("solid", fgColor="FFFFFF")

    hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    dfont  = Font(name='Arial', size=10)
    tfont  = Font(name='Arial', bold=True, color='FFFFFF', size=13)
    bfont  = Font(name='Arial', bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),  bottom=Side(style='thin', color='E5E7EB')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def title_row(ws, text, fill, ncols, row=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill=fill; c.font=tfont; c.alignment=center; c.border=border
        ws.row_dimensions[row].height = 28

    def header_row(ws, headers, fill, row=2):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill=fill; c.font=hfont; c.alignment=center; c.border=border
        ws.row_dimensions[row].height = 20

    def data_row(ws, values, row, alt=True):
        fill = ALT1 if alt else WHITE
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill=fill; c.font=dfont; c.alignment=left; c.border=border
        ws.row_dimensions[row].height = 18

    total_txns = len(df)
    unique_ack = df['Acknowledgement No.'].nunique() if 'Acknowledgement No.' in df.columns else 0
    unique_acc = df['Account Number'].nunique() if 'Account Number' in df.columns else 0
    api_count  = df['Action Taken by'].astype(str).str.upper().str.contains('API', na=False).sum() if 'Action Taken by' in df.columns else 0
    ccc_count  = total_txns - api_count
    api_pct    = round(api_count/total_txns*100, 2) if total_txns else 0
    layer0     = (df['Layers']==0).sum()
    layer1_3   = df['Layers'].between(1,3).sum()
    layer4_5   = df['Layers'].between(4,5).sum()
    layer6_10  = df['Layers'].between(6,10).sum()
    layer10p   = (df['Layers']>10).sum()
    dis_lt500  = (df['Disputed Amount']<500).sum()
    txn_lt500  = (df['Transaction Amount']<500).sum()
    txn_20kp   = (df['Transaction Amount']>=20000).sum()

    # Sheet 1 - Summary
    ws1 = wb.active; ws1.title = "Daily Summary"
    title_row(ws1, f'NCRP DAILY ANALYSIS — {date_label} | Total: {total_txns:,} transactions', NAVY, 4)
    ws1.merge_cells('A3:B3')
    c=ws1.cell(row=3,column=1,value="METRIC"); c.fill=NAVY; c.font=hfont; c.alignment=center; c.border=border
    ws1.merge_cells('C3:D3')
    c=ws1.cell(row=3,column=3,value="VALUE"); c.fill=NAVY; c.font=hfont; c.alignment=center; c.border=border

    metrics = [
        ("Total Transactions", total_txns, BLUE),
        ("Unique ACK Numbers", unique_ack, GREEN),
        ("Unique Account Numbers", unique_acc, PURPLE),
        ("API Attended (SBI API)", api_count, ORANGE),
        ("Unattended (Blank Action Taken by)", int(df["Action Taken by"].isna().sum() + (df["Action Taken by"].astype(str).str.strip() == "").sum()), RED),
        ("CCC", ccc_count, BLUE),
        ("API %", f"{round(api_count/total_txns*100,2)}%", GREEN),
        ("Layer 0 Transactions", layer0, NAVY),
        ("Layer 1-3 Transactions", layer1_3, BLUE),
        ("Layer 4-5 Transactions", layer4_5, ORANGE),
        ("Layer 6-10 Transactions", layer6_10, RED),
        ("Layer 10+ Transactions", layer10p, RED),
        ("Disputed Amount < ₹500", dis_lt500, GREEN),
        ("Disputed Amount ≥ ₹20,000", int((df['Disputed Amount']>=20000).sum()), ORANGE),
        ("Disputed Amount — Layer 1", int(df[df['Layers']==1]['Disputed Amount'].sum()), BLUE),
        ("Disputed Amount — Layer 2", int(df[df['Layers']==2]['Disputed Amount'].sum()), BLUE),
        ("Disputed Amount — Layer 3", int(df[df['Layers']==3]['Disputed Amount'].sum()), BLUE),
        ("Disputed Amount — Layer 1+2+3 Total", int(df[df['Layers'].isin([1,2,3])]['Disputed Amount'].sum()), PURPLE),
        ("Transaction Amount < ₹500", txn_lt500, GREEN),
        ("Transaction Amount ≥ ₹20,000", txn_20kp, ORANGE),
        ("API Attended", int(df[df['Action Taken by'].notna() & (df['Action Taken by'].astype(str).str.strip() != '')].shape[0]), GREEN),
        ("API Unattended (Blank Action Taken by)", int(df[df['Action Taken by'].isna() | (df['Action Taken by'].astype(str).str.strip() == '')].shape[0]), RED),
    ]
    for i, (label, val, fill) in enumerate(metrics, 4):
        ws1.merge_cells(start_row=i,start_column=1,end_row=i,end_column=2)
        c1=ws1.cell(row=i,column=1,value=label)
        c1.fill=ALT1 if i%2==0 else WHITE; c1.font=bfont; c1.alignment=left; c1.border=border
        ws1.merge_cells(start_row=i,start_column=3,end_row=i,end_column=4)
        c2=ws1.cell(row=i,column=3,value=val)
        c2.fill=fill; c2.font=Font(name='Arial',bold=True,color='FFFFFF',size=12)
        c2.alignment=center; c2.border=border
        ws1.row_dimensions[i].height=22
    for col,w in zip(['A','B','C','D'],[30,10,20,10]):
        ws1.column_dimensions[col].width=w

    def breakdown_sheet(ws, title, data, col_headers, fill, col_widths):
        title_row(ws, title, fill, len(col_headers))
        header_row(ws, col_headers, fill)
        for i, (_, row) in enumerate(data.iterrows(), 3):
            vals = [row[c] for c in data.columns]
            data_row(ws, vals, i, i%2==0)
        for col, w in zip([get_column_letter(i+1) for i in range(len(col_widths))], col_widths):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A3'

    # Action
    if 'Action' in df.columns:
        ac = df['Action'].value_counts().reset_index(); ac.columns=['Action','Count']
        ac['% of Total'] = (ac['Count']/total_txns*100).round(2).astype(str)+'%'
        ws2 = wb.create_sheet("Action Breakdown")
        breakdown_sheet(ws2,'ACTION BREAKDOWN',ac,['Action','Count','% of Total'],BLUE,[40,12,14])

    # Layer
    layers_data = pd.DataFrame({
        'Layer Category':['Layer 0','Layer 1-3','Layer 4-5','Layer 6-10','Layer 10+'],
        'Count':[layer0,layer1_3,layer4_5,layer6_10,layer10p]
    })
    layers_data['% of Total'] = (layers_data['Count']/total_txns*100).round(2).astype(str)+'%'
    ws3 = wb.create_sheet("Layer Breakdown")
    breakdown_sheet(ws3,'LAYER-WISE BREAKDOWN',layers_data,['Layer Category','Count','% of Total'],PURPLE,[20,12,14])

    # State
    if 'State' in df.columns:
        sc = df['State'].value_counts().reset_index(); sc.columns=['State','Count']
        sc['% of Total'] = (sc['Count']/total_txns*100).round(2).astype(str)+'%'
        ws4 = wb.create_sheet("State Breakdown")
        breakdown_sheet(ws4,'STATE-WISE BREAKDOWN',sc,['State','Count','% of Total'],GREEN,[35,12,14])

    # Bank
    if 'Money transfer TO Bank' in df.columns:
        bc = df['Money transfer TO Bank'].value_counts().reset_index(); bc.columns=['Bank','Count']
        bc['% of Total'] = (bc['Count']/total_txns*100).round(2).astype(str)+'%'
        ws5 = wb.create_sheet("Bank Breakdown")
        breakdown_sheet(ws5,'MONEY TRANSFER TO BANK',bc,['Bank','Count','% of Total'],ORANGE,[60,12,14])

    # Officer
    if 'Action Taken by' in df.columns:
        oc = df['Action Taken by'].value_counts().reset_index(); oc.columns=['Officer','Count']
        oc['% of Total'] = (oc['Count']/total_txns*100).round(2).astype(str)+'%'
        ws6 = wb.create_sheet("Officer Breakdown")
        breakdown_sheet(ws6,'OFFICER BREAKDOWN',oc,['Officer','Count','% of Total'],NAVY,[35,12,14])

    # Unique ACK
    if 'Acknowledgement No.' in df.columns:
        ack = df.groupby('Acknowledgement No.').agg(
            Transactions=('Acknowledgement No.','count'),
            Total_Amount=('Transaction Amount','sum'),
            Disputed_Amount=('Disputed Amount','sum'),
            Layers=('Layers','first'),
            Mode=('Mode of Payment','first') if 'Mode of Payment' in df.columns else ('Layers','first'),
            State=('State','first') if 'State' in df.columns else ('Layers','first'),
            Action=('Action','first') if 'Action' in df.columns else ('Layers','first'),
        ).reset_index()
        ws7 = wb.create_sheet("Unique ACK List")
        title_row(ws7,f'UNIQUE ACK NUMBERS — {unique_ack:,}',RED,8)
        header_row(ws7,['ACK No.','Transactions','Total Amount','Disputed Amount','Layers','Mode','State','Action'],RED)
        for i,(_, row) in enumerate(ack.iterrows(),3):
            data_row(ws7,[row['Acknowledgement No.'],row['Transactions'],row['Total_Amount'],
                          row['Disputed_Amount'],row['Layers'],row.get('Mode',''),
                          row.get('State',''),row.get('Action','')],i,i%2==0)
        for col,w in zip(['A','B','C','D','E','F','G','H'],[18,14,16,16,8,12,25,30]):
            ws7.column_dimensions[col].width=w
        ws7.freeze_panes='A3'

    # Unique Account
    if 'Account Number' in df.columns:
        acc = df.groupby('Account Number').agg(
            Transactions=('Account Number','count'),
            Total_Amount=('Transaction Amount','sum'),
            Disputed_Amount=('Disputed Amount','sum'),
            Unique_ACKs=('Acknowledgement No.','nunique') if 'Acknowledgement No.' in df.columns else ('Account Number','count'),
            State=('State','first') if 'State' in df.columns else ('Account Number','count'),
            Action=('Action','first') if 'Action' in df.columns else ('Account Number','count'),
        ).reset_index()
        ws8 = wb.create_sheet("Unique Account List")
        title_row(ws8,f'UNIQUE ACCOUNT NUMBERS — {unique_acc:,}',PURPLE,7)
        header_row(ws8,['Account Number','Transactions','Total Amount','Disputed Amount','Unique ACKs','State','Action'],PURPLE)
        for i,(_, row) in enumerate(acc.iterrows(),3):
            data_row(ws8,[row['Account Number'],row['Transactions'],row['Total_Amount'],
                          row['Disputed_Amount'],row.get('Unique_ACKs',''),
                          row.get('State',''),row.get('Action','')],i,i%2==0)
        for col,w in zip(['A','B','C','D','E','F','G'],[22,14,16,16,12,25,30]):
            ws8.column_dimensions[col].width=w
        ws8.freeze_panes='A3'

    # Raw Data
    ws9 = wb.create_sheet("Raw Data")
    title_row(ws9,'RAW TRANSACTION DATA',NAVY,len(df.columns))
    for col,h in enumerate(df.columns,1):
        c=ws9.cell(row=2,column=col,value=h)
        c.fill=NAVY; c.font=hfont; c.alignment=center; c.border=border
    for i,(_, row) in enumerate(df.iterrows(),3):
        fill=ALT1 if i%2==0 else WHITE
        for col,val in enumerate(row,1):
            c=ws9.cell(row=i,column=col,value=val)
            c.fill=fill; c.font=dfont; c.alignment=left; c.border=border
        ws9.row_dimensions[i].height=16
    for col in range(1,len(df.columns)+1):
        ws9.column_dimensions[get_column_letter(col)].width=20
    ws9.freeze_panes='A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ---- MAIN UI ----
uploaded = st.file_uploader("📂 Upload daily NCRP Excel file (.xlsx)", type=["xlsx"])

if uploaded is None:
    st.info("👆 Upload the daily NCRP Excel file to get started")
    st.markdown("### 📋 Expected Format")
    st.markdown("The file should have a sheet with these columns:")
    sample = pd.DataFrame({
        'Acknowledgement No.': ['20107260000485','20107260000485'],
        'Transaction Amount':  [98000, 90000],
        'Disputed Amount':     [0, 0],
        'Layers':              [0, 0],
        'Action':              ['Money transfer TO','Money transfer TO'],
        'State':               ['Delhi','Mumbai'],
        'Account Number':      ['ptminj-1@oksbi','ptminj-1@oksbi'],
        'Mode of Payment':     ['DEB','DEB'],
    })
    st.dataframe(sample, use_container_width=True)

else:
    file_bytes = uploaded.read()
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    sheets = xl.sheet_names

    selected_sheet = st.selectbox("📄 Select sheet with transaction data", sheets,
        index=sheets.index('ViewTransactions') if 'ViewTransactions' in sheets else 0)

    date_label = st.text_input("📅 Enter date label for report", value="Daily Analysis")

    with st.spinner("📊 Analyzing transactions..."):
        df = load_data(file_bytes, selected_sheet)

    if df.empty:
        st.error("❌ No data found in selected sheet")
    else:
        total_txns = len(df)
        unique_ack = df['Acknowledgement No.'].nunique() if 'Acknowledgement No.' in df.columns else 0
        unique_acc = df['Account Number'].nunique() if 'Account Number' in df.columns else 0
        api_count  = df['Action Taken by'].astype(str).str.upper().str.contains('API', na=False).sum() if 'Action Taken by' in df.columns else 0
        ccc_count  = total_txns - api_count
        unattended_n = int(df['Action Taken by'].isna().sum() + (df['Action Taken by'].astype(str).str.strip().isin(['','nan','NaN','None'])).sum()) if 'Action Taken by' in df.columns else 0
        api_pct    = round(api_count/total_txns*100, 2) if total_txns else 0
        layer0     = (df['Layers']==0).sum()
        layer1_3   = df['Layers'].between(1,3).sum()
        layer4_5   = df['Layers'].between(4,5).sum()
        layer6_10  = df['Layers'].between(6,10).sum()
        layer10p   = (df['Layers']>10).sum()

        # Top metrics
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        with c1: st.markdown(f'<div class="metric-card"><div class="metric-num">{total_txns:,}</div><div class="metric-label">Total Transactions</div></div>', unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="metric-card"><div class="metric-num">{unique_ack:,}</div><div class="metric-label">Unique ACKs</div></div>', unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="metric-card"><div class="metric-num">{unique_acc:,}</div><div class="metric-label">Unique Accounts</div></div>', unsafe_allow_html=True)
        with c4: st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#60a5fa">{api_count:,}</div><div class="metric-label">API Attended</div></div>', unsafe_allow_html=True)
        with c5: st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#f87171">{unattended_n:,}</div><div class="metric-label">Unattended</div></div>', unsafe_allow_html=True)
        with c6: st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#34d399">{api_pct}%</div><div class="metric-label">API %</div></div>', unsafe_allow_html=True)
        st.markdown("---")
        tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8 = st.tabs([
            "📊 Summary","🔢 ACK & Accounts","🧅 Layers","⚡ Actions","🗺️ States","🏦 Banks","👮 Officers","🎯 API Analysis"
        ])

        with tab1:
            st.markdown('<div class="section-header">Layer Distribution</div>', unsafe_allow_html=True)
            layer_df = pd.DataFrame({
                'Layer': ['Layer 0','Layer 1-3','Layer 4-5','Layer 6-10','Layer 10+'],
                'Count': [layer0, layer1_3, layer4_5, layer6_10, layer10p]
            })
            col_a, col_b = st.columns(2)
            with col_a:
                fig = px.bar(layer_df, x='Layer', y='Count', color='Layer',
                    color_discrete_sequence=['#2563a8','#059669','#D97706','#DC2626','#7C3AED'],
                    title='Transactions by Layer')
                fig.update_layout(showlegend=False, height=300, paper_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig, use_container_width=True)
            with col_b:
                fig2 = px.pie(layer_df, values='Count', names='Layer', hole=0.4,
                    title='Layer Distribution %',
                    color_discrete_sequence=['#2563a8','#059669','#D97706','#DC2626','#7C3AED'])
                fig2.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown('<div class="section-header">Amount Analysis</div>', unsafe_allow_html=True)
            a1,a2,a3 = st.columns(3)
            a1.metric("Total Transactions", f"{total_txns:,}")
            a2.metric("✅ Attended by API", f"{api_count:,}", f"{api_pct}%")
            a3.metric("❌ Unattended", f"{unattended_n:,}", f"{round(unattended_n/total_txns*100,2)}%")

            st.markdown('<div class="section-header">Disputed Amount ≥ ₹20,000 by Layer</div>', unsafe_allow_html=True)
            l1_dis = int(((df['Layers']==1) & (df['Disputed Amount']>=20000)).sum())
            l2_dis = int(((df['Layers']==2) & (df['Disputed Amount']>=20000)).sum())
            l3_dis = int(((df['Layers']==3) & (df['Disputed Amount']>=20000)).sum())
            l123_dis = l1_dis + l2_dis + l3_dis
            d1, d2, d3, d4 = st.columns(4)
            d1.metric("Layer 1 — Disputed ≥ ₹20k", f"{l1_dis:,}")
            d2.metric("Layer 2 — Disputed ≥ ₹20k", f"{l2_dis:,}")
            d3.metric("Layer 3 — Disputed ≥ ₹20k", f"{l3_dis:,}")
            d4.metric("Layer 1+2+3 Total", f"{l123_dis:,}")

            st.markdown('<div class="section-header">Other Metrics</div>', unsafe_allow_html=True)
            e1, e2, e3 = st.columns(3)
            e1.metric("Disputed Amount < ₹500", f"{(df['Disputed Amount']<500).sum():,}")
            e2.metric("Transaction Amount < ₹500", f"{(df['Transaction Amount']<500).sum():,}")
            e3.metric("Transaction Amount ≥ ₹20,000", f"{(df['Transaction Amount']>=20000).sum():,}")
            a2.metric("Disputed Amount ≥ ₹20,000", f"{(df['Disputed Amount']>=20000).sum():,}")
            a3.metric("Transaction Amount ≥ ₹20,000", f"{(df['Transaction Amount']>=20000).sum():,}")

            st.markdown('<div class="section-header">Disputed Amount by Layer</div>', unsafe_allow_html=True)
            b1, b2, b3, b4 = st.columns(4)
            b1.metric("Layer 1 Disputed", f"₹{int(df[df['Layers']==1]['Disputed Amount'].sum()):,}")
            b2.metric("Layer 2 Disputed", f"₹{int(df[df['Layers']==2]['Disputed Amount'].sum()):,}")
            b3.metric("Layer 3 Disputed", f"₹{int(df[df['Layers']==3]['Disputed Amount'].sum()):,}")
            b4.metric("Layer 1-3 Total", f"₹{int(df[df['Layers'].isin([1,2,3])]['Disputed Amount'].sum()):,}")

            st.markdown('<div class="section-header">API Attended vs Unattended</div>', unsafe_allow_html=True)
            attended = df[df['Action Taken by'].notna() & (df['Action Taken by'].astype(str).str.strip() != '')].shape[0]
            unattended = df[df['Action Taken by'].isna() | (df['Action Taken by'].astype(str).str.strip() == '')].shape[0]
            att_pct = round(attended/total_txns*100, 2)
            unatt_pct = round(unattended/total_txns*100, 2)
            d1, d2, d3 = st.columns(3)
            d1.metric("✅ Attended", f"{attended:,}", f"{att_pct}%")
            d2.metric("❌ Unattended", f"{unattended:,}", f"{unatt_pct}%")
            d3.metric("Transaction Amount < ₹500", f"{(df['Transaction Amount']<500).sum():,}")

        with tab2:
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown('<div class="section-header">Unique ACK Numbers</div>', unsafe_allow_html=True)
                if 'Acknowledgement No.' in df.columns:
                    ack_df = df.groupby('Acknowledgement No.').agg(
                        Transactions=('Acknowledgement No.','count'),
                        Total_Amount=('Transaction Amount','sum'),
                        Disputed_Amount=('Disputed Amount','sum'),
                        Layers=('Layers','first'),
                    ).reset_index()
                    st.metric("Total Unique ACKs", f"{len(ack_df):,}")
                    st.dataframe(ack_df, use_container_width=True, height=350)
            with col_b:
                st.markdown('<div class="section-header">Unique Account Numbers</div>', unsafe_allow_html=True)
                if 'Account Number' in df.columns:
                    acc_df = df.groupby('Account Number').agg(
                        Transactions=('Account Number','count'),
                        Total_Amount=('Transaction Amount','sum'),
                        Disputed_Amount=('Disputed Amount','sum'),
                        Unique_ACKs=('Acknowledgement No.','nunique') if 'Acknowledgement No.' in df.columns else ('Account Number','count'),
                    ).reset_index()
                    st.metric("Total Unique Accounts", f"{len(acc_df):,}")
                    st.dataframe(acc_df, use_container_width=True, height=350)

        with tab3:
            st.markdown('<div class="section-header">Layer-wise Breakdown</div>', unsafe_allow_html=True)
            layer_detail = pd.DataFrame({
                'Layer Category': ['Layer 0','Layer 1-3','Layer 4-5','Layer 6-10','Layer 10+'],
                'Count':          [layer0, layer1_3, layer4_5, layer6_10, layer10p],
                '% of Total':     [f"{round(x/total_txns*100,2)}%" for x in [layer0,layer1_3,layer4_5,layer6_10,layer10p]]
            })
            st.dataframe(layer_detail, use_container_width=True)

            fig3 = px.bar(layer_detail, x='Layer Category', y='Count',
                text='% of Total', color='Layer Category',
                color_discrete_sequence=['#2563a8','#059669','#D97706','#DC2626','#7C3AED'])
            fig3.update_layout(showlegend=False, height=300, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
            st.plotly_chart(fig3, use_container_width=True)

            st.markdown('<div class="section-header">API (Money Transfer TO) per Layer</div>', unsafe_allow_html=True)
            if 'Action' in df.columns:
                def layer_cat(x):
                    if x == 0: return 'Layer 0'
                    elif x <= 3: return 'Layer 1-3'
                    elif x <= 5: return 'Layer 4-5'
                    elif x <= 10: return 'Layer 6-10'
                    else: return 'Layer 10+'
                df['Layer_Cat'] = df['Layers'].apply(layer_cat)

                api_layer = df[df['Action'].str.strip() == 'Money transfer TO'].groupby('Layer_Cat').size().reset_index()
                api_layer.columns = ['Layer Category', 'API Count']
                order = ['Layer 0','Layer 1-3','Layer 4-5','Layer 6-10','Layer 10+']
                api_layer['Layer Category'] = pd.Categorical(api_layer['Layer Category'], categories=order, ordered=True)
                api_layer = api_layer.sort_values('Layer Category')
                api_layer['% of API Total'] = (api_layer['API Count'] / api_layer['API Count'].sum() * 100).round(2).astype(str) + '%'

                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(api_layer, use_container_width=True)
                with col_b:
                    fig_api = px.bar(api_layer, x='Layer Category', y='API Count',
                        text='% of API Total', color='Layer Category',
                        title='API Transactions by Layer',
                        color_discrete_sequence=['#2563a8','#059669','#D97706','#DC2626','#7C3AED'])
                    fig_api.update_layout(showlegend=False, height=300, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                    st.plotly_chart(fig_api, use_container_width=True)

            st.markdown('<div class="section-header">Full Action Breakdown per Layer</div>', unsafe_allow_html=True)
            if 'Action' in df.columns:
                key_actions = ['Money transfer TO', 'Transaction put on hold', 'Others [ <=500]', 'Old Transaction', 'Withdrawal through ATM']
                df['Action_Group'] = df['Action'].apply(lambda x: str(x).strip() if pd.notna(x) and str(x).strip() in key_actions else 'Other')
                crosstab = pd.crosstab(df['Layer_Cat'], df['Action_Group'])
                crosstab = crosstab.reindex(order).fillna(0).astype(int)
                crosstab['Total'] = crosstab.sum(axis=1)
                crosstab = crosstab.reset_index()
                crosstab.columns.name = None
                st.dataframe(crosstab, use_container_width=True)

                # Stacked bar
                melt_df = crosstab.melt(id_vars='Layer_Cat', value_vars=key_actions + ['Other'],
                    var_name='Action', value_name='Count')
                fig_stack = px.bar(melt_df, x='Layer_Cat', y='Count', color='Action',
                    title='Action Breakdown by Layer (Stacked)',
                    barmode='stack',
                    color_discrete_sequence=['#2563a8','#DC2626','#D97706','#059669','#7C3AED','#6B7280'])
                fig_stack.update_layout(height=380, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                st.plotly_chart(fig_stack, use_container_width=True)

        with tab4:
            st.markdown('<div class="section-header">Action Breakdown</div>', unsafe_allow_html=True)
            if 'Action' in df.columns:
                ac = df['Action'].value_counts().reset_index(); ac.columns=['Action','Count']
                ac['% of Total'] = (ac['Count']/total_txns*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(ac, use_container_width=True, height=350)
                with col_b:
                    fig4 = px.pie(ac, values='Count', names='Action', hole=0.4,
                        title='Action Distribution')
                    fig4.update_layout(height=350, paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig4, use_container_width=True)

        with tab5:
            st.markdown('<div class="section-header">State-wise Breakdown</div>', unsafe_allow_html=True)
            if 'State' in df.columns:
                sc = df['State'].value_counts().reset_index(); sc.columns=['State','Count']
                sc['% of Total'] = (sc['Count']/total_txns*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(sc, use_container_width=True, height=400)
                with col_b:
                    fig5 = px.bar(sc.head(15), x='Count', y='State', orientation='h',
                        title='Top 15 States', color='Count',
                        color_continuous_scale=['#EFF6FF','#2563a8'])
                    fig5.update_layout(height=400, paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig5, use_container_width=True)

        with tab6:
            st.markdown('<div class="section-header">Bank Breakdown</div>', unsafe_allow_html=True)
            if 'Money transfer TO Bank' in df.columns:
                bc = df['Money transfer TO Bank'].value_counts().reset_index(); bc.columns=['Bank','Count']
                bc['% of Total'] = (bc['Count']/total_txns*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(bc, use_container_width=True, height=400)
                with col_b:
                    fig6 = px.bar(bc.head(10), x='Count', y='Bank', orientation='h',
                        title='Top 10 Banks', color='Count',
                        color_continuous_scale=['#EFF6FF','#D97706'])
                    fig6.update_layout(height=400, paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig6, use_container_width=True)

        with tab7:
            st.markdown('<div class="section-header">Officer-wise Action Breakdown</div>', unsafe_allow_html=True)
            if 'Action Taken by' in df.columns:
                oc = df['Action Taken by'].value_counts().reset_index(); oc.columns=['Officer','Count']
                oc['% of Total'] = (oc['Count']/total_txns*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(oc, use_container_width=True, height=400)
                with col_b:
                    fig7 = px.bar(oc.head(15), x='Count', y='Officer', orientation='h',
                        title='Officer Performance', color='Count',
                        color_continuous_scale=['#EFF6FF','#7C3AED'])
                    fig7.update_layout(height=400, paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig7, use_container_width=True)

        with tab8:
            st.markdown('<div class="section-header">API vs Put on Hold vs Others</div>', unsafe_allow_html=True)
            df['Action_clean'] = df['Action'].astype(str).str.strip()
            df['Category'] = df['Action_clean'].apply(
                lambda x: 'API (Money Transfer TO)' if x == 'Money transfer TO'
                else 'Put on Hold' if x == 'Transaction put on hold'
                else 'Others'
            )
            cat_df = df['Category'].value_counts().reset_index()
            cat_df.columns = ['Category','Count']
            cat_df['% of Total'] = (cat_df['Count']/total_txns*100).round(2).astype(str)+'%'

            c1, c2, c3 = st.columns(3)
            api_n  = cat_df[cat_df['Category']=='API (Money Transfer TO)']['Count'].values[0]
            hold_n = cat_df[cat_df['Category']=='Put on Hold']['Count'].values[0]
            oth_n  = cat_df[cat_df['Category']=='Others']['Count'].values[0]
            c1.metric("🔵 API (Money Transfer TO)", f"{api_n:,}", f"{round(api_n/total_txns*100,2)}%")
            c2.metric("🔴 Put on Hold", f"{hold_n:,}", f"{round(hold_n/total_txns*100,2)}%")
            c3.metric("⚪ Others", f"{oth_n:,}", f"{round(oth_n/total_txns*100,2)}%")

            col_a, col_b = st.columns(2)
            with col_a:
                st.dataframe(cat_df, use_container_width=True)
            with col_b:
                fig_cat = px.pie(cat_df, values='Count', names='Category', hole=0.4,
                    title='API vs Hold vs Others',
                    color_discrete_sequence=['#2563a8','#DC2626','#6B7280'])
                fig_cat.update_layout(height=300, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                st.plotly_chart(fig_cat, use_container_width=True)

            api_df = df[df['Action_clean'] == 'Money transfer TO'].copy()

            st.markdown('<div class="section-header">API by Layer</div>', unsafe_allow_html=True)
            def layer_cat(x):
                if x == 0: return 'Layer 0'
                elif x <= 3: return 'Layer 1-3'
                elif x <= 5: return 'Layer 4-5'
                elif x <= 10: return 'Layer 6-10'
                else: return 'Layer 10+'
            api_df['Layer_Cat'] = api_df['Layers'].apply(layer_cat)
            api_layer = api_df['Layer_Cat'].value_counts().reset_index()
            api_layer.columns = ['Layer','API Count']
            order = ['Layer 0','Layer 1-3','Layer 4-5','Layer 6-10','Layer 10+']
            api_layer['Layer'] = pd.Categorical(api_layer['Layer'], categories=order, ordered=True)
            api_layer = api_layer.sort_values('Layer')
            api_layer['% of API'] = (api_layer['API Count']/api_n*100).round(2).astype(str)+'%'

            col_a, col_b = st.columns(2)
            with col_a:
                st.dataframe(api_layer, use_container_width=True)
            with col_b:
                fig_al = px.bar(api_layer, x='Layer', y='API Count', text='% of API',
                    color='Layer', title='API Transactions by Layer',
                    color_discrete_sequence=['#2563a8','#059669','#D97706','#DC2626','#7C3AED'])
                fig_al.update_layout(showlegend=False, height=300, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                st.plotly_chart(fig_al, use_container_width=True)

            st.markdown('<div class="section-header">API by State</div>', unsafe_allow_html=True)
            if 'State' in api_df.columns:
                api_state = api_df['State'].value_counts().reset_index()
                api_state.columns = ['State','API Count']
                api_state['% of API'] = (api_state['API Count']/api_n*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(api_state, use_container_width=True, height=350)
                with col_b:
                    fig_as = px.bar(api_state.head(12), x='API Count', y='State', orientation='h',
                        title='Top States — API Transactions', color='API Count',
                        color_continuous_scale=['#EFF6FF','#2563a8'])
                    fig_as.update_layout(height=380, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                    st.plotly_chart(fig_as, use_container_width=True)

            st.markdown('<div class="section-header">API by Bank</div>', unsafe_allow_html=True)
            if 'Money transfer TO Bank' in api_df.columns:
                api_bank = api_df['Money transfer TO Bank'].value_counts().reset_index()
                api_bank.columns = ['Bank','API Count']
                api_bank['% of API'] = (api_bank['API Count']/api_n*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(api_bank, use_container_width=True, height=350)
                with col_b:
                    fig_ab = px.bar(api_bank.head(10), x='API Count', y='Bank', orientation='h',
                        title='Top Banks — API Transactions', color='API Count',
                        color_continuous_scale=['#FFF7ED','#D97706'])
                    fig_ab.update_layout(height=380, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                    st.plotly_chart(fig_ab, use_container_width=True)

            st.markdown('<div class="section-header">API by Officer</div>', unsafe_allow_html=True)
            if 'Action Taken by' in api_df.columns:
                api_officer = api_df['Action Taken by'].value_counts().reset_index()
                api_officer.columns = ['Officer','API Count']
                api_officer['% of API'] = (api_officer['API Count']/api_n*100).round(2).astype(str)+'%'
                col_a, col_b = st.columns(2)
                with col_a:
                    st.dataframe(api_officer, use_container_width=True, height=350)
                with col_b:
                    fig_ao = px.bar(api_officer.head(12), x='API Count', y='Officer', orientation='h',
                        title='Top Officers — API Transactions', color='API Count',
                        color_continuous_scale=['#F5F3FF','#7C3AED'])
                    fig_ao.update_layout(height=380, paper_bgcolor='rgba(0,0,0,0)', font_color='white')
                    st.plotly_chart(fig_ao, use_container_width=True)

            st.markdown("---")
            st.markdown('<div class="section-header">Download API-only Report</div>', unsafe_allow_html=True)
            api_excel = io.BytesIO()
            with pd.ExcelWriter(api_excel, engine='openpyxl') as writer:
                cat_df.to_excel(writer, sheet_name='API vs Hold vs Others', index=False)
                api_layer.to_excel(writer, sheet_name='API by Layer', index=False)
                if 'State' in api_df.columns:
                    api_state.to_excel(writer, sheet_name='API by State', index=False)
                if 'Money transfer TO Bank' in api_df.columns:
                    api_bank.to_excel(writer, sheet_name='API by Bank', index=False)
                if 'Action Taken by' in api_df.columns:
                    api_officer.to_excel(writer, sheet_name='API by Officer', index=False)
                api_df.drop(columns=['Action_clean','Category','Layer_Cat'], errors='ignore').to_excel(
                    writer, sheet_name='Raw API Transactions', index=False)
            api_excel.seek(0)
            st.download_button(
                "📥 Download API Analysis Report",
                api_excel,
                "API_Analysis_Report.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )           

        st.markdown("---")
        with st.spinner("📥 Preparing Excel report..."):
            excel_buf = build_excel(df, date_label)
        st.download_button(
            "📥 Download Full Excel Report",
            excel_buf,
            f"NCRP_Analysis_{date_label.replace(' ','_')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
